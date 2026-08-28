import frappe
import csv
import openpyxl
import io
from openpyxl.styles import Font, PatternFill, Alignment
from frappe.utils import nowdate

'''
所在的表单:
    Amazon Order Summary == 基础订单
    Amazon Order Import Session == 基础订单的主表导入的表
    Amazon internal form binding table title == 对应翻译字段
    Import Amazon and remove orders == 导入移除订单页面
    Import Amazon order settlement details == 导入结算订单页面
    Amazon Settlement Detail == 亚马逊结算明细存储表
'''



@frappe.whitelist()
def 检查是否绑定了标题(站点区,所在的表单):
    #开始分配规则,开始
    match 所在的表单:
        #基础订单
        case "Amazon Order Summary":
            # 采集所有的翻译对照表,开始
            filters = {
                "亚马逊区域": 站点区,     # 替换为你的实际字段名
                "单据类型": 所在的表单        # 替换为你的实际字段名
            }
            # 2. 执行获取操作
            翻译对照表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters, 
                fields=["外部表格标题", "翻译成中文"]
            )

            if not 翻译对照表:
                frappe.throw(f"此站点（{站点区}）未配置【{所在的表单}】的绑定表格标题！")
            return True
        case "Amazon removes order details":
            # 采集所有的翻译对照表,开始
            filters = {
                "亚马逊区域": 站点区,     # 替换为你的实际字段名
                "单据类型": 所在的表单        # 替换为你的实际字段名
            }
            # 2. 执行获取操作
            翻译对照表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters, 
                fields=["外部表格标题", "翻译成中文"]
            )
            if not 翻译对照表:
                frappe.throw(f"此站点（{站点区}）未配置【{所在的表单}】的绑定表格标题！")
            return True
        case "Amazon Settlement Detail":
            # 采集所有的翻译对照表,开始
            filters = {
                "亚马逊国家": 站点区,     # 替换为你的实际字段名
                "单据类型": 所在的表单        # 替换为你的实际字段名
            }
            # 2. 执行获取操作
            翻译对照表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters, 
                fields=["外部表格标题", "翻译成中文"]
            )
            if not 翻译对照表:
                frappe.throw(f"此站点（{站点区}）未配置【{所在的表单}】的绑定表格标题！")
            return True


def 翻译亚马逊表格(需要翻译的文件,所在的表单,self):

    需要染成绿色 = ["已发货", "shipped", "是", "true", "seiiable", "可销售的", "完成的", "标准"]
    需要染成红色 = ["取消了", "cancelled", "canceled", "否", "false", "disposal", "处理", "未发货", "加急"]

    #开始分配规则,开始
    match 所在的表单:
        #基础订单
        case "Amazon Order Summary":



            filters = {
                "亚马逊区域": self.站点区,     # 替换为你的实际字段名
                "单据类型": 所在的表单        # 替换为你的实际字段名
            }
            # 2. 执行获取操作
            翻译对照表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters, 
                fields=["外部表格标题", "翻译成中文"]
            )


            翻译字典 = {}
            for b in 翻译对照表:
                英文标题 = str(b.外部表格标题).strip().lower()
                中文标题 = b.翻译成中文
                翻译字典[英文标题] = 中文标题
            # 采集所有的翻译对照表,结束

            # 1. 自动判断编码 (一行流)
            编码 = 'utf-8-sig'
            try: open(需要翻译的文件, 'r', encoding=编码).read(4096)
            except: 编码 = 'gbk'

            # 确定编码后开始循环
            翻译后的订单列表 = []
            with open(需要翻译的文件, 'r', encoding=编码) as f_in:
                样本 = f_in.read(4096)
                f_in.seek(0)
                # 谁的数量多,确定格式
                tab个数 = 样本.count('\t')
                逗号个数 = 样本.count(',')
                分隔符 = '\t' if tab个数 > 逗号个数 else ','
                # 它会自动把第一行撕下来当成 Key (键)
                读取器 = csv.DictReader(f_in, delimiter=分隔符)
                # 拿到所有的原始英文标题
                标题 = 读取器.fieldnames  # 例如: ['amazon-order-id', 'sku', ...]
                # 挨个打印一行一行的数据
                # 1. 定义两组不同的关键列
                # 【复制组】：前面有带括号的，后面原始位置也保留
                复制组 = ['sku', 'quantity', 'purchase-date', 'amazon-order-id']

                # 【搬家组】：前面有带括号的，后面原始位置彻底删除
                搬家组 = [
                    'ship-promotion-discount', 'item-promotion-discount', 'gift-wrap-tax',
                    'gift-wrap-price', 'shipping-tax', 'shipping-price', 'item-tax', 'item-price'
                ]

                # A 阶段处理这两组的总和
                所有前置列 = 复制组 + 搬家组

                for 行数据 in 读取器:
                    翻译后的行 = {}
                    
                    # --- A. 先把这两组都“复制”一份到前面 (带方括号) ---
                    for 关键名 in 所有前置列:
                        原始值 = 行数据.get(关键名, "")
                        匹配名 = 关键名.strip().lower()
                        
                        # 强制数字转换逻辑（针对数量和所有金额列）
                        处理后的值 = 原始值
                        if 匹配名 == 'quantity' or 匹配名 in 搬家组:
                            try:
                                if str(原始值).strip():
                                    数字值 = float(str(原始值).strip())
                                    处理后的值 = int(数字值) if 数字值 == int(数字值) else 数字值
                            except: pass

                        中文名 = 翻译字典.get(匹配名, 关键名)
                        翻译后的行[f"[{中文名}]"] = 处理后的值

                    # --- B. 正常的循环翻译 ---
                    for 英文键, 值 in 行数据.items():
                        匹配键 = str(英文键).strip().lower()
                        
                        # ✨ 关键：只有在“搬家组”里的列才跳过！
                        # 基础组（sku, quantity等）不跳过，所以后面还会再出一遍，实现“复制”
                        if 匹配键 in 搬家组:
                            continue
                            
                        中文键 = 翻译字典.get(匹配键, 英文键)
                        
                        # 正常的翻译和反查转数字逻辑
                        匹配值 = str(值).strip().lower() if 值 else ""
                        中文值 = 翻译字典.get(匹配值, 值)
                        
                        if 中文值 == 值:
                            try:
                                数字值 = float(str(中文值).strip())
                                中文值 = int(数字值) if 数字值 == int(数字值) else 数字值
                            except: pass
                                
                        翻译后的行[中文键] = 中文值

                    翻译后的订单列表.append(翻译后的行)
            
         


            #写入到xlsx内
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "翻译后的基础订单"
            if 翻译后的订单列表:
                表头 = list(翻译后的订单列表[0].keys())
                
                ws.append(表头)
                # 填充颜色：浅蓝色
                表头填充 = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                # 字体：加粗
                表头字体 = Font(bold=True, size=12)
                # 居中对齐
                居中 = Alignment(horizontal="center", vertical="center")
                for cell in ws[1]:  # 循环第一行的每一个单元格
                    cell.fill = 表头填充
                    cell.font = 表头字体
                    cell.alignment = 居中
                # 1. 先定义好黄色背景样式（放在循环外面，省内存）
                黄色背景 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                绿色 = "008000"
                红色 = "FF0000"
                # 2. 填充内容行
                for 行索引, 行数据 in enumerate(翻译后的订单列表, start=2):
                    内容行 = [行数据.get(列名, "") for 列名 in 表头]
                    ws.append(内容行)
                    # 拿到刚刚 append 进去的那一行（也就是当前行）
                    当前行单元格 = ws[ws.max_row] 
                    for cell in 当前行单元格:
                        cell.alignment = 居中
                        # --- ✨ 关键修改：如果是前四列（A, B, C, D），背景涂黄 ---
                        if cell.column <= 4:
                            cell.fill = 黄色背景
                        # --- 状态文字颜色逻辑（红绿灯） ---
                        单元格内容 = str(cell.value).strip().lower() if cell.value else ""
                        # 绿色加粗逻辑
                        if 单元格内容 in 需要染成绿色:
                            cell.font = Font(color=绿色, bold=True)
                        # 红色加粗逻辑
                        if 单元格内容 in 需要染成红色:
                            cell.font = Font(color=红色, bold=True)
                # 4. 自动调整列宽（让表格看起来更舒服）
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # 获取列字母
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width


            # 5. 保存文件
            # 1. 将 Excel 写入内存流而不是硬盘
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            # 获取当前日期
            日期 = nowdate() 
            # 拼接一个更详细的文件名
            frappe.response['filename'] = f"翻译亚马逊基础订单_{self.name}_{日期}.xlsx"
            frappe.response['filecontent'] = output.getvalue()
            frappe.response['type'] = 'binary'
        case "Amazon removes order details":


            filters = {
                "亚马逊区域": self.项目类型_选择区域,     # 替换为你的实际字段名
                "单据类型": 所在的表单        # 替换为你的实际字段名
            }
            # 2. 执行获取操作
            翻译对照表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters, 
                fields=["外部表格标题", "翻译成中文"]
            )

            翻译字典 = {}
            for b in 翻译对照表:
                英文标题 = str(b.外部表格标题).strip().lower()
                中文标题 = b.翻译成中文
                翻译字典[英文标题] = 中文标题



            # 1. 自动判断编码 (一行流)
            编码 = 'utf-8-sig'
            try: open(需要翻译的文件, 'r', encoding=编码).read(4096)
            except: 编码 = 'gbk'

            # 确定编码后开始循环
            翻译后的订单列表 = []
            with open(需要翻译的文件, 'r', encoding=编码) as f_in:
                样本 = f_in.read(4096)
                f_in.seek(0)
                # 谁的数量多,确定格式
                tab个数 = 样本.count('\t')
                逗号个数 = 样本.count(',')
                分隔符 = '\t' if tab个数 > 逗号个数 else ','
                # 它会自动把第一行撕下来当成 Key (键)
                读取器 = csv.DictReader(f_in, delimiter=分隔符)
                # 1. 获取映射关系 (这里已经包含了你所有的翻译规则)


            
                for 行数据 in 读取器:
                    翻译后的行 = {}
                    # --- B. 正常的循环翻译 ---
                    for 英文键, 值 in 行数据.items():
                        匹配键 = str(英文键).strip().lower()

                        中文键 = 翻译字典.get(匹配键, 英文键)
                        
                        # 正常的翻译和反查转数字逻辑
                        匹配值 = str(值).strip().lower() if 值 else ""
                        中文值 = 翻译字典.get(匹配值, 值)

                        if 中文值 == 值:
                            try:
                                数字值 = float(str(中文值).strip())
                                中文值 = int(数字值) if 数字值 == int(数字值) else 数字值
                            except: pass
                                
                        翻译后的行[中文键] = 中文值

                    翻译后的订单列表.append(翻译后的行)
            


            #写入到xlsx内
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "翻译后的移除订单"
            if 翻译后的订单列表:
                表头 = list(翻译后的订单列表[0].keys())
                
                ws.append(表头)
                # 填充颜色：浅蓝色
                表头填充 = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                # 字体：加粗
                表头字体 = Font(bold=True, size=12)
                # 居中对齐
                居中 = Alignment(horizontal="center", vertical="center")
                for cell in ws[1]:  # 循环第一行的每一个单元格
                    cell.fill = 表头填充
                    cell.font = 表头字体
                    cell.alignment = 居中
                # 1. 先定义好黄色背景样式（放在循环外面，省内存）
                绿色 = "008000"
                红色 = "FF0000"
                # 2. 填充内容行
                for 行索引, 行数据 in enumerate(翻译后的订单列表, start=2):
                    内容行 = [行数据.get(列名, "") for 列名 in 表头]
                    ws.append(内容行)
                    # 拿到刚刚 append 进去的那一行（也就是当前行）
                    当前行单元格 = ws[ws.max_row] 
                    for cell in 当前行单元格:
                        cell.alignment = 居中
                        # --- 状态文字颜色逻辑（红绿灯） ---
                        单元格内容 = str(cell.value).strip().lower() if cell.value else ""
                        # 绿色加粗逻辑
                        if 单元格内容 in 需要染成绿色:
                            cell.font = Font(color=绿色, bold=True)
                        # 红色加粗逻辑
                        if 单元格内容 in 需要染成红色:
                            cell.font = Font(color=红色, bold=True)
                # 4. 自动调整列宽（让表格看起来更舒服）
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # 获取列字母
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width


            # 5. 保存文件
            # 1. 将 Excel 写入内存流而不是硬盘
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            # 获取当前日期
            日期 = nowdate()
            # 拼接一个更详细的文件名
            frappe.response['filename'] = f"翻译亚马逊移除订单_{self.name}_{日期}.xlsx"
            frappe.response['filecontent'] = output.getvalue()
            frappe.response['type'] = 'binary'
        case "Amazon Settlement Detail":


            filters = {
                "亚马逊国家": self.店铺国家,     # 替换为你的实际字段名
                "单据类型": 所在的表单        # 替换为你的实际字段名
            }
            # 2. 执行获取操作
            翻译对照表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters, 
                fields=["外部表格标题", "翻译成中文"]
            )
            print(self.店铺国家)
            print(所在的表单)
            print(翻译对照表)
            翻译字典 = {}
            for b in 翻译对照表:
                英文标题 = str(b.外部表格标题).strip().lower()
                中文标题 = b.翻译成中文
                翻译字典[英文标题] = 中文标题
            print(翻译字典)
            # 1. 自动判断编码 (一行流)
            编码 = 'utf-8-sig'
            try: open(需要翻译的文件, 'r', encoding=编码).read(4096)
            except: 编码 = 'gbk'

            # 确定编码后开始循环
            描述列表 = []
            翻译后的订单列表 = []
            with open(需要翻译的文件, 'r', encoding=编码) as f_in:
                样本 = f_in.read(4096)
                f_in.seek(0)
                # 谁的数量多,确定格式
                tab个数 = 样本.count('\t')
                逗号个数 = 样本.count(',')
                分隔符 = '\t' if tab个数 > 逗号个数 else ','
                #其他情况是数据或者标题
                # 我们要找包含 "settlement id" 或 "date/time" 的那一行
                普通读取器 = csv.reader(f_in, delimiter=分隔符)
                跳过行数 = 0
                for 行列表 in 普通读取器:
                    # 清理空格，判断这行是不是我们要的表头
                    有效列 = [格.strip() for 格 in 行列表 if 格.strip()]
                    # 亚马逊表头通常很长（>20列），描述行很短（1列）
                    if len(有效列) < 5: 
                        # 获取原始的这一行文字（取出第一个元素并转小写）
                        # 即使这一行有多个空格，我们也取第一个非空内容来匹配
                        英文文本 = 有效列[0].lower() if 有效列 else ""
                        # 3. 去翻译字典里找，找不到就用原话
                        # 注意：翻译字典里的 Key 建议预先存为小写，匹配更准
                        中文描述 = 翻译字典.get(英文文本, 英文文本)
                        描述列表.append(中文描述)
                        跳过行数 += 1
                    else:
                        # 找到了疑似表头的一行！停止普通读取
                        break
                #找全部的数据
                f_in.seek(0)
                # 彻底跳过刚才那些描述行
                for _ in range(跳过行数):
                    f_in.readline()
                # 现在的第一行就是真正的表头了
                读取器 = csv.DictReader(f_in, delimiter=分隔符)
                优先列 = ["product sales", "product sales tax", "shipping credits", "shipping credits tax", "gift wrap credits", "giftwrap credits tax", "Regulatory Fee", "Tax On Regulatory Fee", "promotional rebates", "promotional rebates tax", "marketplace withheld tax", "selling fees", "fba fees", "other transaction fees", "other", "total"]
                for 行数据 in 读取器:
                    # 此时的 行数据 就是干净的字典：{"date/time": "...", "settlement id": "..."}
                    翻译后的行 = {}
                    已处理英文键 = set()

                    def 处理值逻辑(值, 翻译字典):
                        if 值 is None: return ""
                        匹配值 = str(值).strip().lower()
                        中文值 = 翻译字典.get(匹配值, 值)

                        # 如果没翻译成功，尝试转数字
                        if 中文值 == 值:
                            try:
                                # 清理千分位逗号（如果有的话，比如 1,234.56）
                                清理后的值 = str(值).replace(',', '').strip()
                                数字值 = float(清理后的值)
                                return int(数字值) if 数字值 == int(数字值) else 数字值
                            except:
                                pass
                        return 中文值

                    # --- 第一步：先处理优先列 ---
                    # 我们遍历一遍行数据，先把匹配上优先列的找出来
                    for 英文键, 值 in 行数据.items():
                        匹配键 = str(英文键).strip().lower()
                        if 匹配键 in 优先列:
                            中文键 = 翻译字典.get(匹配键, 英文键)
                            # 执行你的翻译和转数字逻辑 (封装成一个小逻辑)
                            处理后的值 = 处理值逻辑(值, 翻译字典)
                            翻译后的行[中文键] = 处理后的值
                            已处理英文键.add(英文键)

                    # --- 第二步：处理剩下的所有列 ---
                    for 英文键, 值 in 行数据.items():
                        if 英文键 in 已处理英文键:
                            continue
                            
                        匹配键 = str(英文键).strip().lower()
                        中文键 = 翻译字典.get(匹配键, 英文键)
                        处理后的值 = 处理值逻辑(值, 翻译字典)
                        
                        翻译后的行[中文键] = 处理后的值

                    翻译后的订单列表.append(翻译后的行)

            #写入到xlsx内
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "翻译后的结算订单"
            if 翻译后的订单列表:
                #写入描述
                for 行索引, 行数据 in enumerate(描述列表):
                    ws.append([行数据])
                #写入表头
                表头 = list(翻译后的订单列表[0].keys())
                
                ws.append(表头)
                表头行号 = ws.max_row
                # 填充颜色：浅蓝色
                表头填充 = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                # 字体：加粗
                表头字体 = Font(bold=True, size=12)
                # 居中对齐
                居中 = Alignment(horizontal="center", vertical="center")
                for cell in ws[表头行号]:  # 循环第一行的每一个单元格
                    cell.fill = 表头填充
                    cell.font = 表头字体
                    cell.alignment = 居中
                # 1. 先定义好黄色背景样式（放在循环外面，省内存）
                绿色 = "008000"
                红色 = "FF0000"
                # 2. 填充内容行
                for 行索引, 行数据 in enumerate(翻译后的订单列表, start=2):
                    内容行 = [行数据.get(列名, "") for 列名 in 表头]
                    ws.append(内容行)
                    # 拿到刚刚 append 进去的那一行（也就是当前行）
                    当前行单元格 = ws[ws.max_row] 
                    for cell in 当前行单元格:
                        cell.alignment = 居中
                        # --- 状态文字颜色逻辑（红绿灯） ---
                        单元格内容 = str(cell.value).strip().lower() if cell.value else ""
                        # 绿色加粗逻辑
                        if 单元格内容 in 需要染成绿色:
                            cell.font = Font(color=绿色, bold=True)
                        # 红色加粗逻辑
                        if 单元格内容 in 需要染成红色:
                            cell.font = Font(color=红色, bold=True)
                # 4. 自动调整列宽（让表格看起来更舒服）
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # 获取列字母
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

            # 5. 保存文件
            # 1. 将 Excel 写入内存流而不是硬盘
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            # 获取当前日期
            日期 = nowdate()
            # 拼接一个更详细的文件名
            frappe.response['filename'] = f"翻译亚马逊结算订单_{self.name}_{日期}.xlsx"
            frappe.response['filecontent'] = output.getvalue()
            frappe.response['type'] = 'binary'
        case _:
            pass

def 亚马逊文件关联表存储表是否一致(订单原始文件路径,所在的表单,站点区域和国家):
    match 所在的表单:
        #基础订单
        case "Amazon Order Summary":
            # 采集所有的对应表,开始
            filters = {
                "单据类型": "Amazon Order Summary",
                "选择表": "基础订单",
                "亚马逊区域": 站点区域和国家  # 新增的条件
            }

            # 2. 执行获取操作
            配置映射表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters,
                fields=["外部表格标题", "erpnext内部字段"]
            )

            # 3. 校验结果：如果没有找到任何配置，直接弹窗报错并拦截
            if not 配置映射表:
                # 1. 这种方式只弹窗，不回滚数据库，不强行报错
                frappe.msgprint(
                    msg=f"站点【{站点区域和国家}】尚未配置【Amazon Order Summary - 基础订单】的映射关系，请先去配置。",
                    title="配置缺失",
                    indicator="orange" # 橘色警告，比红色报错温和
                )
                
                # 2. 【关键】手动停止后续代码执行
                return
            
            字段字典 = {}
            for b in 配置映射表:
                内部字段 = str(b.erpnext内部字段).strip().lower()
                外部表格标题 = b.外部表格标题
                字段字典[内部字段] = 外部表格标题
            # 采集所有的对应表,结束

            #拿合法字段
            合法字段 = frappe.get_meta("Amazon Order Summary").get_valid_columns()

            #拿文件标题
            # 1. 自动判断编码 (一行流)
            编码 = 'utf-8-sig'
            try: open(订单原始文件路径, 'r', encoding=编码).read(4096)
            except: 编码 = 'gbk'
            # 确定编码后开始循环
            原始标题列表 = []
            with open(订单原始文件路径, 'r', encoding=编码) as f_in:
                样本 = f_in.read(4096)
                f_in.seek(0)
                # 谁的数量多,确定格式
                tab个数 = 样本.count('\t')
                逗号个数 = 样本.count(',')
                分隔符 = '\t' if tab个数 > 逗号个数 else ','
                # 它会自动把第一行撕下来当成 Key (键)
                读取器 = csv.DictReader(f_in, delimiter=分隔符)
                # 拿到所有的原始英文标题
                if 读取器.fieldnames:
                    原始标题列表 = [str(h).strip() for h in 读取器.fieldnames]
                else:
                    原始标题列表 = []


            # 准备三项标题对比
            配置缺失标题 = []
            字段非法标题 = []
            # 2. 遍历文件里的每一个标题
            for 单个外部标题 in 原始标题列表:
                找到对应内部字段 = None
                
                # 💡 核心：在你的【字段字典】里反向查找
                # 遍历字典，看哪一个“值”匹配当前的“文件标题”
                for 内部字段, 外部表格标题 in 字段字典.items():
                    if str(外部表格标题 or "").lower() == str(单个外部标题 or "").lower():
                        找到对应内部字段 = 内部字段
                        break 
                
                # 3. 开始三段式逻辑校验
                if not 找到对应内部字段:
                    # 情况 A：文件里有这一列，但你的“配置映射表”里没定义对应的外部标题
                    配置缺失标题.append(单个外部标题)
                else:
                    # 情况 B：配置里有，但我们要检查这个“内部字段”在 ERP 数据库里真不真实
                    # 排除掉你那些“纯翻译词汇”（比如你定义的翻译词可能不在合法列里）
                    if 找到对应内部字段 not in 合法字段:
                        字段非法标题.append(f"{单个外部标题} -> {找到对应内部字段}")
            # 4. 汇总报错并弹窗
            错误汇总 = []
            if 配置缺失标题:
                错误汇总.append(f"<b>以下列在配置表中未定义：</b><br>{' , '.join(配置缺失标题)}")
            if 字段非法标题:
                错误汇总.append(f"<b>以下映射的内部字段在 ERP 中不存在：</b><br>{' , '.join(字段非法标题)}")

            if 错误汇总:
                # 使用 frappe.throw 强制停止并弹窗
                frappe.throw("<br><br>".join(错误汇总), title="字段一致性检查失败")


            文件字段映射字典 = {}

            for 单个外部标题 in 原始标题列表:
                # 再次反向查找（因为前面已经校验过，这次一定能找到）
                for 内部字段, 外部表格标题 in 字段字典.items():
                    if str(外部表格标题 or "").lower() == str(单个外部标题 or "").lower():
                        文件字段映射字典[单个外部标题] = 内部字段
                        break

            # 6. 返回结果
            # 如果是在函数内，直接 return
            return 文件字段映射字典
        case "Amazon removes order details":
            # 采集所有的对应表,开始
            # 配置映射表 = frappe.get_all("Amazon internal form binding table title", filters={"单据类型": "Amazon removes order details","选择表": "移除订单"},fields=["外部表格标题","erpnext内部字段"] )
            # 字段字典 = {}






            # 采集所有的对应表,开始
            filters = {
                "单据类型": "Amazon removes order details",
                "选择表": "移除订单",
                "亚马逊区域": 站点区域和国家  # 新增的条件
            }

            # 2. 执行获取操作
            配置映射表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters,
                fields=["外部表格标题", "erpnext内部字段"]
            )
            # 3. 校验结果：如果没有找到任何配置，直接弹窗报错并拦截
            if not 配置映射表:
                # 1. 这种方式只弹窗，不回滚数据库，不强行报错
                frappe.msgprint(
                    msg=f"站点【{站点区域和国家}】尚未配置【Amazon removes order details - 基础订单】的映射关系，请先去配置。",
                    title="配置缺失",
                    indicator="orange" # 橘色警告，比红色报错温和
                )
                
                # 2. 【关键】手动停止后续代码执行
                return







            字段字典 = {}
            for b in 配置映射表:
                内部字段 = str(b.erpnext内部字段).strip().lower()
                外部表格标题 = b.外部表格标题
                字段字典[内部字段] = 外部表格标题
            # 采集所有的对应表,结束
            合法字段 = frappe.get_meta("Amazon removes order details").get_valid_columns()
            #拿文件标题
            # 1. 自动判断编码 (一行流)
            编码 = 'utf-8-sig'
            try: open(订单原始文件路径, 'r', encoding=编码).read(4096)
            except: 编码 = 'gbk'
            # 确定编码后开始循环
            原始标题列表 = []
            with open(订单原始文件路径, 'r', encoding=编码) as f_in:
                样本 = f_in.read(4096)
                f_in.seek(0)
                # 谁的数量多,确定格式
                tab个数 = 样本.count('\t')
                逗号个数 = 样本.count(',')
                分隔符 = '\t' if tab个数 > 逗号个数 else ','
                # 它会自动把第一行撕下来当成 Key (键)
                读取器 = csv.DictReader(f_in, delimiter=分隔符)
                # 拿到所有的原始英文标题
                if 读取器.fieldnames:
                    原始标题列表 = [str(h).strip() for h in 读取器.fieldnames]
                else:
                    原始标题列表 = []


            # 准备三项标题对比
            配置缺失标题 = []
            字段非法标题 = []
            # 2. 遍历文件里的每一个标题
            for 单个外部标题 in 原始标题列表:
                找到对应内部字段 = None
                
                # 💡 核心：在你的【字段字典】里反向查找
                # 遍历字典，看哪一个“值”匹配当前的“文件标题”
                for 内部字段, 外部表格标题 in 字段字典.items():
                    if str(外部表格标题 or "").lower() == str(单个外部标题 or "").lower():
                        找到对应内部字段 = 内部字段
                        break 

                # 3. 开始三段式逻辑校验
                if not 找到对应内部字段:
                    # 情况 A：文件里有这一列，但你的“配置映射表”里没定义对应的外部标题
                    配置缺失标题.append(单个外部标题)
                else:
                    # 情况 B：配置里有，但我们要检查这个“内部字段”在 ERP 数据库里真不真实
                    # 排除掉你那些“纯翻译词汇”（比如你定义的翻译词可能不在合法列里）
                    if 找到对应内部字段 not in 合法字段:
                        字段非法标题.append(f"{单个外部标题} -> {找到对应内部字段}")
            # 4. 汇总报错并弹窗
            错误汇总 = []
            if 配置缺失标题:
                错误汇总.append(f"<b>以下列在配置表中未定义：</b><br>{' , '.join(配置缺失标题)}")
            if 字段非法标题:
                错误汇总.append(f"<b>以下映射的内部字段在 ERP 中不存在：</b><br>{' , '.join(字段非法标题)}")

            if 错误汇总:
                # 使用 frappe.throw 强制停止并弹窗
                frappe.throw("<br><br>".join(错误汇总), title="字段一致性检查失败")


            文件字段映射字典 = {}

            for 单个外部标题 in 原始标题列表:
                # 再次反向查找（因为前面已经校验过，这次一定能找到）
                for 内部字段, 外部表格标题 in 字段字典.items():
                    if str(外部表格标题 or "").lower() == str(单个外部标题 or "").lower():
                        文件字段映射字典[单个外部标题] = 内部字段
                        break

            # 6. 返回结果
            # 如果是在函数内，直接 return
            return 文件字段映射字典
        case "Amazon Settlement Detail":
            # 采集所有的对应表,开始
            #配置映射表 = frappe.get_all("Amazon internal form binding table title", filters={"单据类型": "Amazon Settlement Detail","选择表": "结算订单"},fields=["外部表格标题","erpnext内部字段"] )
            




            # 采集所有的对应表,开始
            filters = {
                "单据类型": "Amazon Settlement Detail",
                "选择表": "结算订单",
                "亚马逊国家": 站点区域和国家  # 新增的条件
            }

            # 2. 执行获取操作
            配置映射表 = frappe.get_all(
                "Amazon internal form binding table title", 
                filters=filters,
                fields=["外部表格标题", "erpnext内部字段"]
            )
            # 3. 校验结果：如果没有找到任何配置，直接弹窗报错并拦截
            if not 配置映射表:
                # 1. 这种方式只弹窗，不回滚数据库，不强行报错
                frappe.msgprint(
                    msg=f"站点【{站点区域和国家}】尚未配置【Amazon Settlement Detail - 基础订单】的映射关系，请先去配置。",
                    title="配置缺失",
                    indicator="orange" # 橘色警告，比红色报错温和
                )
                
                # 2. 【关键】手动停止后续代码执行
                return






            字段字典 = {}
            for b in 配置映射表:
                内部字段 = str(b.erpnext内部字段).strip().lower()
                外部表格标题 = b.外部表格标题
                字段字典[内部字段] = 外部表格标题
            # 采集所有的对应表,结束
            合法字段 = frappe.get_meta("Amazon Settlement Detail").get_valid_columns()
            #拿文件标题
            # 1. 自动判断编码 (一行流)
            编码 = 'utf-8-sig'
            try: open(订单原始文件路径, 'r', encoding=编码).read(4096)
            except: 编码 = 'gbk'
            # 确定编码后开始循环
            原始标题列表 = []
            with open(订单原始文件路径, 'r', encoding=编码) as f_in:
                样本 = f_in.read(4096)
                f_in.seek(0)
                # 谁的数量多,确定格式
                tab个数 = 样本.count('\t')
                逗号个数 = 样本.count(',')
                分隔符 = '\t' if tab个数 > 逗号个数 else ','
                # 它会自动把第一行撕下来当成 Key (键)
                读取器 = csv.reader(f_in, delimiter=分隔符)
                跳过行数 = 0
                for 行列表 in 读取器:
                    # 清理空格，判断这行是不是我们要的表头
                    有效列 = [格.strip() for 格 in 行列表 if 格.strip()]
                    # 亚马逊表头通常很长（>20列），描述行很短（1列）
                    if len(有效列) < 5: 
                        pass
                        跳过行数 += 1
                    else:
                        # 找到了疑似表头的一行！停止普通读取
                        break

                #找全部的数据
                f_in.seek(0)
                # 彻底跳过刚才那些描述行
                for _ in range(跳过行数):
                    f_in.readline()
                # 现在的第一行就是真正的表头了
                读取器 = csv.DictReader(f_in, delimiter=分隔符)
                # 1. 在循环外先拿好标题
                if 读取器.fieldnames:
                    原始标题列表 = [str(h).strip() for h in 读取器.fieldnames if h]
                else:
                    原始标题列表 = []
            


            # 准备三项标题对比
            配置缺失标题 = []
            字段非法标题 = []
            # 2. 遍历文件里的每一个标题
            for 单个外部标题 in 原始标题列表:
                找到对应内部字段 = None
                
                # 💡 核心：在你的【字段字典】里反向查找
                # 遍历字典，看哪一个“值”匹配当前的“文件标题”
                for 内部字段, 外部表格标题 in 字段字典.items():
                    # 将两个变量都转为小写后再对比
                    if str(外部表格标题 or "").lower() == str(单个外部标题 or "").lower():
                        找到对应内部字段 = 内部字段
                        break 

                # 3. 开始三段式逻辑校验
                if not 找到对应内部字段:
                    # 情况 A：文件里有这一列，但你的“配置映射表”里没定义对应的外部标题
                    配置缺失标题.append(单个外部标题)
                else:
                    # 情况 B：配置里有，但我们要检查这个“内部字段”在 ERP 数据库里真不真实
                    # 排除掉你那些“纯翻译词汇”（比如你定义的翻译词可能不在合法列里）
                    if 找到对应内部字段 not in 合法字段:
                        字段非法标题.append(f"{单个外部标题} -> {找到对应内部字段}")
            # 4. 汇总报错并弹窗
            错误汇总 = []
            if 配置缺失标题:
                list_html = "".join([f"<li>{item}</li>" for item in 配置缺失标题])
                错误汇总.append(f"<b>以下列在配置表中未定义：</b><ul style='margin-top: 5px;'>{list_html}</ul>")
            if 字段非法标题:
                错误汇总.append(f"<b>以下映射的内部字段在 ERP 中不存在：</b><br>{' , '.join(字段非法标题)}")

            if 错误汇总:
                # 使用 frappe.throw 强制停止并弹窗
                frappe.throw("<br><br>".join(错误汇总), title="字段一致性检查失败")


            文件字段映射字典 = {}

            for 单个外部标题 in 原始标题列表:
                # 再次反向查找（因为前面已经校验过，这次一定能找到）
                for 内部字段, 外部表格标题 in 字段字典.items():
                    if str(外部表格标题 or "").lower() == str(单个外部标题 or "").lower():
                        文件字段映射字典[单个外部标题] = 内部字段
                        break

            # 6. 返回结果
            # 如果是在函数内，直接 return
            return 文件字段映射字典
        case _:
            pass